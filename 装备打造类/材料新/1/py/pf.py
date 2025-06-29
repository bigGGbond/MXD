import openpyxl
import json
from collections import defaultdict
import sys

# 设置编码
sys.stdout.reconfigure(encoding='utf-8')

# 打开 Excel 文件
try:
    wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sc\scripts\材料新\excel\披风.xlsx', data_only=True)
    ws = wb['Sheet1']  # 使用第一个工作表
    print(f"成功打开工作簿，工作表名称: {ws.title}")
except Exception as e:
    print(f"打开Excel文件时出错: {str(e)}")
    sys.exit(1)

# 打印表头信息
print("表头信息：")
for cell in ws[1]:
    if cell.value:
        print(cell.value)

def null_if_none(value):
    """将None值转换为null"""
    return value if value is not None else None

def process_excel():
    # 初始化数据结构
    data = {}

    # 遍历每一行
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:  # 跳过表头
            continue

        # 检查行是否为空或数据不完整
        if not row or len(row) < 15:  # 确保至少有所需的所有列
            print(f"警告：第 {row_idx + 1} 行数据不完整或为空: {row}")
            continue

        try:
            # 获取基本数据
            cape_id = str(row[0])      # 披风ID
            level = int(row[1])        # 等级
            gong_mo = row[2]           # 攻/魔
            fang_mo_fang = row[3]      # 防/魔防
            tiao_yue = row[4]          # 跳跃力

            # 材料数据
            boss_material = {
                'id': null_if_none(row[5]),    # BOSS材料
                'bind': null_if_none(row[6]),  # BOSS绑定
                'qty': row[7] or 0             # BOSS数量
            }

            material1 = {
                'id': null_if_none(row[8]),    # 材料1
                'bind': null_if_none(row[9]),  # 材料1绑定
                'qty': row[10] or 0            # 材料1数量
            }

            material2 = {
                'id': null_if_none(row[11]),   # 材料2
                'bind': null_if_none(row[12]), # 材料2绑定
                'qty': row[13] or 0            # 材料2数量
            }

            gold = row[14] or 0               # 金币

            # 如果是新的披风ID，初始化数据结构
            if cape_id not in data:
                data[cape_id] = {}

            # 构造等级数据
            level_data = {
                "双攻": gong_mo,
                "双防": fang_mo_fang,
                "跳跃": tiao_yue,
                "材料": [
                    [boss_material['id'], boss_material['qty']],
                    [material1['id'], material1['qty']],
                    [material2['id'], material2['qty']],
                    [0, gold]
                ],
                "绑定材料": [
                    boss_material['bind'],
                    material1['bind'],
                    material2['bind']
                ]
            }

            # 添加到数据结构
            data[cape_id][f"等级{level}"] = level_data

        except Exception as e:
            print(f"处理第 {row_idx + 1} 行时发生错误: {str(e)}")
            continue

    return data

def save_json(data, filename='披风奖励数据.json'):
    """保存为JSON文件"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def main():
    try:
        print("开始处理Excel数据...")
        data = process_excel()
        print("数据处理完成，正在保存...")
        save_json(data)
        print("文件保存成功！")
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    main()
