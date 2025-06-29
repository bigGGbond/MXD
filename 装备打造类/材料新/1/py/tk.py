import openpyxl
import json
from collections import defaultdict

# 打开 Excel 文件
wb = openpyxl.load_workbook(r'C:\Users\李柱柱\Desktop\scripts\头盔.xlsx')
ws = wb['Sheet1']

# 打印表头信息
print("表头信息：")
for cell in ws[1]:
    print(cell.value)

# 第一遍收集所有头盔ID与阶数
item_records = defaultdict(dict)

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue
    item_id = str(row[0])
    step = int(row[1])
    item_records[item_id] = {
        "阶数": step,
        "原始行": row
    }

# 构建 阶数 => 头盔ID 映射
step_to_id_map = {}
for item_id, info in item_records.items():
    step_to_id_map[info["阶数"]] = item_id

# 用于替换 None 为 null
def null_if_none(value):
    return value if value is not None else None

# 主数据结构
data = {}

# 第二遍处理数据
for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue

    print(f"处理第 {row_idx} 行，列数：{len(row)}")
    print(f"行内容：{row}")

    item_id = str(row[0])            # 当前头盔ID
    current_step = int(row[1])       # 当前阶数
    level = row[2]                   # 等级（1~10）
    gong_mo = row[3]                 # 攻/魔
    fang_mo_fang = row[4]            # 防/魔防

    # 材料 ID 和数量
    boss_id = null_if_none(row[5])   # BOSS材料
    boss_bind = null_if_none(row[6]) # BOSS绑定
    boss_qty = row[7] or 0           # BOSS数量

    material1_id = null_if_none(row[8])   # 材料1
    material1_bind = null_if_none(row[9])  # 材料1绑定
    material1_qty = row[10] or 0     # 材料1数量

    gold_id = 0
    gold_qty = row[11] or 0          # 金币

    # 构造材料列表
    material_list = [
        [boss_id, boss_qty],
        [material1_id, material1_qty],
        [gold_id, gold_qty]
    ]

    # 构造绑定材料列表
    bind_list = [
        boss_bind,       # boss_bind
        material1_bind   # material1_bind
    ]

    # 等级键（保留数字）
    level_key = f"等级{level}"

    # 初始化结构
    if item_id not in data:
        next_step_id = step_to_id_map.get(current_step + 1, 0)
        data[item_id] = {
            "进阶id": int(next_step_id) if next_step_id else 0
        }

    # 更新每个等级的属性
    data[item_id][level_key] = {
        "攻/魔": gong_mo,         # 攻/魔
        "防/魔防": fang_mo_fang, # 防/魔防
        "材料": material_list,
        "绑定材料": bind_list
    }

# 保存为 JSON 文件
with open('头盔奖励数据.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
