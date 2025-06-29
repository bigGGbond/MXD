import openpyxl
import json

# 设置编码
import sys
sys.stdout.reconfigure(encoding='utf-8')

# 打开 Excel 文件
wb = openpyxl.load_workbook(r'C:\Users\李柱柱\Desktop\sc\scripts\材料新\excel\神环.xlsx')
ws = wb['Sheet1']

# 打印表头信息
print("表头信息：")
for cell in ws[1]:
    print(cell.value)

# 用于替换 None 为 null
def null_if_none(value):
    return value if value is not None else None

# 先收集所有数据，按 item_id, 阶数, 等级 分类
raw_data = {}
for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue
    if not row[0]:
        continue
    item_id = str(row[0])
    stage = int(row[1])
    level = int(row[2])
    lmys = row[3]
    fang_mo_fang = row[4]
    boss_id = null_if_none(row[5])
    boss_bind = null_if_none(row[6])
    boss_qty = row[7] or 0
    material1_id = null_if_none(row[8])
    material1_bind = None  # 材料1无绑定
    material1_qty = row[9] or 0
    material2_id = null_if_none(row[10])
    material2_bind = null_if_none(row[11])
    material2_qty = row[12] or 0
    gold_id = 0
    gold_qty = row[13] or 0
    material_list = [
        [boss_id, boss_qty],
        [material1_id, material1_qty],
        [material2_id, material2_qty],
        [gold_id, gold_qty]
    ]
    bind_list = [
        boss_bind,
        material1_bind,
        material2_bind
    ]
    # 存入临时结构
    raw_data.setdefault(item_id, {})
    raw_data[item_id].setdefault(stage, {})
    raw_data[item_id][stage][level] = {
        "力敏运智": lmys,
        "防/魔防": fang_mo_fang,
        "材料": material_list,
        "绑定材料": bind_list
    }

# 构建最终结构
final_data = {}
for item_id in raw_data:
    final_data[item_id] = {}
    for stage in range(1, 14):  # 1~13
        stage_key = f"阶段{stage}"
        final_data[item_id][stage_key] = {"进阶id": int(item_id)}
        for level in range(1, 11):  # 1~10
            level_key = f"等级{level}"
            # 如果有数据则填充，否则不填
            if stage in raw_data[item_id] and level in raw_data[item_id][stage]:
                final_data[item_id][stage_key][level_key] = raw_data[item_id][stage][level]

# 保存为 JSON 文件
with open('神环奖励数据.json', 'w', encoding='utf-8') as f:
    json.dump(final_data, f, ensure_ascii=False, indent=2)

print("\n数据处理完成")
