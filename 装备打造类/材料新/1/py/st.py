import openpyxl
import json
from collections import defaultdict
import sys

# ���ñ���
sys.stdout.reconfigure(encoding='utf-8')

# �� Excel �ļ�
try:
    wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sc\scripts\������\excel\����.xlsx', data_only=True)
    ws = wb['Sheet1']  # ʹ�õ�һ��������
    print(f"�ɹ��򿪹�����������������: {ws.title}")
except Exception as e:
    print(f"��Excel�ļ�ʱ����: {str(e)}")
    sys.exit(1)

# ��ӡ��ͷ��Ϣ
print("��ͷ��Ϣ��")
for cell in ws[1]:
    if cell.value:
        print(cell.value)

def null_if_none(value):
    """��Noneֵת��Ϊnull"""
    return value if value is not None else None

def process_excel():
    # ��ʼ�����ݽṹ
    data = {}

    # ����ÿһ��
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:  # ������ͷ
            continue

        # ������Ƿ�Ϊ�ջ����ݲ�����
        if not row or len(row) < 15:  # ȷ�������������������
            print(f"���棺�� {row_idx + 1} �����ݲ�������Ϊ��: {row}")
            continue

        try:
            # ��ȡ��������
            cape_id = str(row[0])      # ����ID
            level = int(row[1])        # �ȼ�
            gong_mo = row[2]           # ��/ħ
            fang_mo_fang = row[3]      # ��/ħ��

            # ��������
            boss_material = {
                'id': null_if_none(row[4]),    # BOSS����
                'bind': null_if_none(row[5]),  # BOSS��
                'qty': row[6] or 0             # BOSS����
            }

            material1 = {
                'id': null_if_none(row[7]),    # ����1
                'bind': null_if_none(row[8]),  # ����1��
                'qty': row[9] or 0            # ����1����
            }

            material2 = {
                'id': null_if_none(row[10]),   # ����2
                'bind': null_if_none(row[11]), # ����2��
                'qty': row[12] or 0            # ����2����
            }

            gold = row[13] or 0               # ���

            # ������µ�����ID����ʼ�����ݽṹ
            if cape_id not in data:
                data[cape_id] = {}

            # ����ȼ�����
            level_data = {
                "˫��": gong_mo,
                "˫��": fang_mo_fang,
                "����": [
                    [boss_material['id'], boss_material['qty']],
                    [material1['id'], material1['qty']],
                    [material2['id'], material2['qty']],
                    [0, gold]
                ],
                "�󶨲���": [
                    boss_material['bind'],
                    material1['bind'],
                    material2['bind']
                ]
            }

            # ��ӵ����ݽṹ
            data[cape_id][f"�ȼ�{level}"] = level_data

        except Exception as e:
            print(f"����� {row_idx + 1} ��ʱ��������: {str(e)}")
            continue

    return data

def save_json(data, filename='���׽�������.json'):
    """����ΪJSON�ļ�"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def main():
    try:
        print("��ʼ����Excel����...")
        data = process_excel()
        print("���ݴ�����ɣ����ڱ���...")
        save_json(data)
        print("�ļ�����ɹ���")
    except Exception as e:
        print(f"��������: {str(e)}")

if __name__ == "__main__":
    main()
