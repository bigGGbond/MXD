import openpyxl
import json
from collections import defaultdict

# 打开 Excel 文件
wb = openpyxl.load_workbook(r'C:\Users\李柱柱\Desktop\scripts\材料\项链.xlsx')
ws = wb['Sheet1']

# 第一遍收集所有项链 ID 与阶数
item_records = defaultdict(dict)

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue
    item_id = str(row[0])
    step = int(row[1])
    item_records[item_id] = {
        "阶数": step,
        "原始行": row
    }

# 构建 阶数 => 项链ID 映射
step_to_id_map = {}
for item_id, info in item_records.items():
    step_to_id_map[info["阶数"]] = item_id

# 用于替换 None 为 null
def null_if_none(value):
    return value if value is not None else None

# 主数据结构
data = {}

# 第二遍处理数据
for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue

    item_id = str(row[0])            # 当前项链ID
    current_step = int(row[1])       # 当前阶数
    level = row[2]                   # 等级（1~10）
    li_min_yun_zhi = row[3]          # 四维
    fang_mo_fang = row[4]            # 双防

    # 材料 ID 和数量
    boss_id = null_if_none(row[5])
    boss_bind = null_if_none(row[6])
    boss_qty = row[7] or 0

    xianglian_id = null_if_none(row[8])
    xianglian_bind = null_if_none(row[9])
    xianglian_qty = row[10] or 0

    maple_id = null_if_none(row[11])
    maple_bind = null_if_none(row[12])
    maple_qty = row[13] or 0

    gold_id = 0
    gold_qty = row[14] or 0

    # 构造材料列表
    material_list = [
        [boss_id, boss_qty],
        [xianglian_id, xianglian_qty],
        [maple_id, maple_qty],
        [gold_id, gold_qty]
    ]

    # 构造绑定材料列表
    bind_list = [
        boss_bind,
        xianglian_bind,
        maple_bind
    ]

    # 等级键（保留数字）
    level_key = f"等级{level}"

    # 初始化结构
    if item_id not in data:
        next_step_id = step_to_id_map.get(current_step + 1, 0)
        data[item_id] = {
            "进阶ID": int(next_step_id) if next_step_id else 0
        }

    data[item_id][level_key] = {
        "四维": li_min_yun_zhi,
        "双防": fang_mo_fang,
        "材料": material_list,
        "绑定材料": bind_list
    }

# 保存为 JSON 文件
with open('项链奖励数据.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
