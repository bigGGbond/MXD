import openpyxl
import json
from collections import defaultdict
import sys

# 设置编码
sys.stdout.reconfigure(encoding='utf-8')

# 打开 Excel 文件
try:
    wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\sc\scripts\材料新\披风.xlsx', data_only=True)
    ws = wb[wb.sheetnames[0]]  # 使用第一个工作表
    print(f"成功打开工作簿，工作表名称: {ws.title}")
except Exception as e:
    print(f"打开Excel文件时出错: {str(e)}")
    sys.exit(1)

# 打印表头信息
print("表头信息：")
for cell in ws[1]:
    if cell.value:
        print(cell.value)

def null_if_none(value):
    """处理空值，返回0或原值"""
    return 0 if value is None else value

# 主数据结构
data = {}

# 第一遍收集所有披风ID与阶数
item_records = defaultdict(dict)
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:  # 跳过空行
        continue
    try:
        item_id = str(null_if_none(row[0]))
        step = int(null_if_none(row[1]))
        item_records[item_id] = {
            "阶数": step,
            "原始行": row
        }
    except Exception as e:
        print(f"处理行数据时出错: {str(e)}")
        print(f"行数据: {row}")
        continue

# 构建 阶数 => 披风ID 映射
step_to_id_map = {}
for item_id, info in item_records.items():
    step_to_id_map[info["阶数"]] = item_id

print("开始处理数据...")

# 处理数据
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    try:
        if not row[0]:  # 跳过空行
            continue
            
        print(f"正在处理第 {row_idx} 行")
        print(f"行数据: {row}")
        
        # 基础数据
        item_id = str(null_if_none(row[0]))
        current_step = int(null_if_none(row[1]))
        level = int(null_if_none(row[2]))
        
        # 属性数据
        gong_mo = int(null_if_none(row[3]))
        fang_mo_fang = int(null_if_none(row[4]))
        tiao_yue = int(null_if_none(row[5]))
        
        # 材料数据
        material_list = [
            [null_if_none(row[6]), null_if_none(row[8])],     # BOSS材料及数量
            [null_if_none(row[9]), null_if_none(row[11])],    # 材料1及数量
            [null_if_none(row[12]), null_if_none(row[14])],   # 材料2及数量
            [0, null_if_none(row[15])]                        # 金币数量
        ]
        
        bind_list = [
            null_if_none(row[7]),     # BOSS材料绑定
            null_if_none(row[10]),    # 材料1绑定
            null_if_none(row[13])     # 材料2绑定
        ]
        
        # 初始化物品数据结构
        if item_id not in data:
            next_step_id = step_to_id_map.get(current_step + 1, 0)
            data[item_id] = {
                "进阶id": int(next_step_id) if next_step_id else 0
            }
        
        # 更新等级数据
        data[item_id][f"等级{level}"] = {
            "攻/魔": gong_mo,
            "防/魔防": fang_mo_fang,
            "跳跃力": tiao_yue,
            "材料": material_list,
            "绑定材料": bind_list
        }
        
    except Exception as e:
        print(f"处理第 {row_idx} 行时出错: {str(e)}")
        print(f"行数据: {row}")
        continue

# 保存为 JSON 文件
try:
    with open('披风奖励数据.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("\n数据处理完成")
except Exception as e:
    print(f"保存JSON文件时出错: {str(e)}")
