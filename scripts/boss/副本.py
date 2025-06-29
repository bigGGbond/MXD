import openpyxl
import json

# 材料ID映射，注意这里根据你给的材料名称写全了
材料ID = {
    "金币": 0,
    "副本蛋": 4170005,
    "绑定副本蛋": 4170006,  # 你这个和金币一样？确认一下
    "概率额外1副本蛋": 4170005  # 这个ID你没给，我随便填了个占位符，需换成正确ID
}

filename = r'C:\Users\Admin\Desktop\sc\scripts\boss\副本.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb.active

奖励数据 = {}

# Excel中每个副本表头间有空行隔开，先确定每个副本起始行，按副本名“副本1”“副本2”……来定位
副本行号列表 = []
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
    if row[1] and isinstance(row[1], str) and row[1].startswith("副本"):
        副本行号列表.append(i)

# 每个副本处理
for idx, start_row in enumerate(副本行号列表):
    副本名 = sheet.cell(row=start_row, column=2).value  # B列是副本名

    # 奖励列表初始化
    奖励数据[副本名] = []

    # 数据起始行 = 副本名所在行 + 2 (因为下一行是表头)
    数据起始行 = start_row + 2

    # 该副本数据终止行（不含），是下一个副本名行或者表尾
    终止行 = 副本行号列表[idx + 1] if idx + 1 < len(副本行号列表) else sheet.max_row + 1

    # 遍历该副本所有等级数据行
    for r in range(数据起始行, 终止行):
        row = list(sheet.iter_rows(min_row=r, max_row=r, max_col=6, values_only=True))[0]

        等级 = row[0]
        if 等级 is None:
            continue  # 空行跳过
        try:
            int(等级)
        except:
            continue  # 等级不是数字跳过

        # 取数据，确保float，None时用0
        def to_float(v):
            try:
                return round(float(v), 4)
            except:
                return 0

        金币 = to_float(row[1])
        副本蛋 = to_float(row[2])
        绑定副本蛋 = to_float(row[3])
        概率额外 = to_float(row[4])

        奖励列表 = [
            [材料ID["金币"], 金币],
            [材料ID["副本蛋"], 副本蛋],
            [材料ID["绑定副本蛋"], 绑定副本蛋],
            [材料ID["概率额外1副本蛋"], 概率额外],
        ]

        奖励数据[副本名].append(奖励列表)

# 保存结果
with open("副本.json", "w", encoding="utf-8") as f:
    json.dump(奖励数据, f, ensure_ascii=False, indent=2)

print("数据导出完成！")
