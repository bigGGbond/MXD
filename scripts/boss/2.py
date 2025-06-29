import openpyxl
import json

# 排名段位对应列偏移
排名段位列偏移 = {
    1: 0,    # 第1名
    3: 16,   # 第2-3名
    6: 32,   # 第4-6名
    10: 48,  # 第7-10名
    11: 64   # 第11名
}

wb = openpyxl.load_workbook(r'C:\Users\李柱柱\Desktop\sc\scripts\boss\团队远征.xlsx')
ws = wb.active

# 读取第二行表头
header = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=2, max_row=2))]

# 找"暗影币"起始列
起始索引 = -1
for i, col_name in enumerate(header):
    if col_name == "暗影币":
        起始索引 = i
        break

if 起始索引 == -1:
    raise ValueError("未找到'暗影币'列")

所有奖励数据 = []

# 处理所有数据行
for row in ws.iter_rows(min_row=3, values_only=True):
    if not any(row):  # 跳过空行
        continue
        
    奖励列表 = []
    
    # 处理每个排名段位的奖励
    for 段位, 基础偏移 in 排名段位列偏移.items():
        段位奖励 = [段位]
        
        # 处理8对材料（普通材料和绑定材料）
        for i in range(8):
            col_index = 起始索引 + 基础偏移 + i * 2
            
            # 获取普通材料ID和数量
            材料ID = row[col_index]
            数量 = row[col_index + 1]
            
            if 材料ID is not None and 数量 is not None:
                材料ID = int(材料ID)
                数量 = round(float(数量), 2)
                段位奖励.append([材料ID, 数量])
            else:
                段位奖励.append([0, 0])

        奖励列表.append(段位奖励)
    
    if 奖励列表:  # 如果这行数据有效，添加到总数据中
        所有奖励数据.append(奖励列表)

# 保存为json
with open("团队远征奖励数据.json", "w", encoding="utf-8") as f:
    json.dump(所有奖励数据, f, ensure_ascii=False, indent=2)

print("数据导出完成！")
